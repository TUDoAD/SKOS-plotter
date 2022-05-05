import argparse
from pathlib import Path
from typing import List, Tuple, Dict
from typing import Literal
import logging

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from colorama import init, Fore, Style

from pydantic.error_wrappers import ValidationError
from vocexcel import models
from vocexcel import profiles
import pyshacl
#from vocexcel import __version__

import os

RDF_FILE_ENDINGS = {
    ".ttl": "ttl",
    ".rdf": "xml",
    ".xml": "xml",
    ".json-ld": "json-ld",
    ".json": "json-ld",
    ".nt": "nt",
    ".n3": "n3",
}
EXCEL_FILE_ENDINGS = ["xlsx"]
KNOWN_FILE_ENDINGS = [str(x) for x in RDF_FILE_ENDINGS.keys()] + EXCEL_FILE_ENDINGS
template_version = "0.4.0"


class ConversionError(Exception):
    pass


def split_and_tidy(cell_value: str):
    return (
        [x.strip() for x in cell_value.strip().split(",")]
        if cell_value is not None
        else None
    )


def extract_concepts_and_collections(
    s: Worksheet,
) -> Tuple[List[models.Concept], List[models.Collection]]:
    concepts = []
    collections = []
    process_concept = False
    process_collection = False
    for col in s.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if cell.value == "Concept URI":
                process_concept = True
            elif cell.value == "Collection URI":
                process_concept = False
                process_collection = True
            elif process_concept:
                if cell.value is None:
                    pass
                else:
                    try:
                        global template_version
                        if template_version == "0.3.0":
                            c = models.Concept(
                                uri=s[f"A{row}"].value,
                                pref_label=s[f"B{row}"].value,
                                pl_language_code=split_and_tidy(s[f"C{row}"].value),  # new in 0.3.0
                                alt_labels=split_and_tidy(s[f"D{row}"].value),
                                definition=s[f"E{row}"].value,
                                def_language_code=split_and_tidy(s[f"F{row}"].value),  # new in 0.3.0
                                children=split_and_tidy(s[f"G{row}"].value),
                                other_ids=split_and_tidy(s[f"H{row}"].value),
                                home_vocab_uri=s[f"I{row}"].value,
                                provenance=s[f"J{row}"].value,
                            )
                        elif template_version == "0.2.1":
                            c = models.Concept(
                                uri=s[f"A{row}"].value,
                                pref_label=s[f"B{row}"].value,
                                alt_labels=split_and_tidy(s[f"C{row}"].value),
                                definition=s[f"D{row}"].value,
                                children=split_and_tidy(s[f"E{row}"].value),
                                other_ids=split_and_tidy(s[f"F{row}"].value),
                                home_vocab_uri=s[f"G{row}"].value,
                                provenance=s[f"H{row}"].value,
                            )
                        elif template_version == "0.2.2": #This version was created by NFDI4Cat-TA1 based on template version 0.2.1
                            c = models.Concept(
                                uri=s[f"A{row}"].value,
                                pref_label=s[f"B{row}"].value,
                                alt_labels=split_and_tidy(s[f"C{row}"].value),
                                definition=s[f"D{row}"].value,
                                children=split_and_tidy(s[f"E{row}"].value),
                                related=split_and_tidy(s[f"F{row}"].value),
                                other_ids=split_and_tidy(s[f"G{row}"].value),
                                home_vocab_uri=s[f"H{row}"].value,
                                provenance=s[f"I{row}"].value,
                            )

                        else:
                            raise ConversionError("The version of the Excel template your are using is not recognised")

                        concepts.append(c)
                    except ValidationError as e:
                        raise ConversionError(
                            f"Concept processing error, row {row}, error: {e}"
                        )
            elif process_collection:
                if cell.value is None:
                    pass
                else:
                    try:
                        c = models.Collection(
                            uri=s[f"A{row}"].value,
                            pref_label=s[f"B{row}"].value,
                            definition=s[f"C{row}"].value,
                            members=split_and_tidy(s[f"D{row}"].value),
                            provenance=s[f"E{row}"].value,
                        )
                        collections.append(c)
                    except ValidationError as e:
                        raise ConversionError(
                            f"Collection processing error, row {row}, error: {e}"
                        )
            elif cell.value is None:
                pass

    return concepts, collections


def excel_to_rdf(
    file_to_convert_path: Path,
    sheet_name=None,
    output_type: Literal["file", "string", "graph"] = "file",
    output_file_path=None,
    output_format: Literal["turtle", "xml", "json-ld"] = "turtle",
):
    """Converts a sheet within an Excel workbook to an RDF file"""
    if type(file_to_convert_path) is str:
        file_to_convert_path = Path(file_to_convert_path)
    if not file_to_convert_path.name.endswith(tuple(EXCEL_FILE_ENDINGS)):
        raise ValueError("Files for conversion to RDF must be Excel files ending .xlsx")
    wb = load_workbook(filename=str(file_to_convert_path), data_only=True)
    pi = wb["program info"]
    global template_version
    template_version = pi["B2"].value

    sheet = wb["vocabulary" if sheet_name is None else sheet_name]

    # Vocabulary
    try:
        cs = models.ConceptScheme(
            uri=sheet["B1"].value,
            title=sheet["B2"].value,
            description=sheet["B3"].value,
            created=sheet["B4"].value,
            modified=sheet["B5"].value,
            creator=sheet["B6"].value,
            publisher=sheet["B7"].value,
            version=sheet["B8"].value,
            provenance=sheet["B9"].value,
            custodian=sheet["B10"].value,
            pid=sheet["B11"].value,
        )
    except ValidationError as e:
        raise ConversionError(f"ConceptScheme processing error: {e}")

    # Concepts & Collections
    concepts, collections = extract_concepts_and_collections(sheet)

    # Build the total vocab
    v = models.Vocabulary(concept_scheme=cs, concepts=concepts, collections=collections)

    # Write out the file
    if output_type == "graph":
        return v.to_graph()
    elif output_type == "string":
        return v.to_graph().serialize(format=output_format)
    else:  # output_format == "file":
        if output_file_path is not None:
            #dest = output_file_path
            dest = Path(output_file_path + os.path.basename(file_to_convert_path)).with_suffix(".ttl")
        else:
            if output_format == "xml":
                suffix = ".rdf"
            elif output_format == "json-ld":
                suffix = ".json-ld"
            else:
                suffix = ".ttl"
            dest = file_to_convert_path.with_suffix(suffix)
        v.to_graph().serialize(destination=str(dest), format=output_format)
        return dest


def rdf_to_excel(
        file_to_convert_path: Path,
        profile="vocpub",
        output_file_path=None,
        error_level=1,
        message_level=1,
        log_file=None
):
    if type(file_to_convert_path) is str:
        file_to_convert_path = Path(file_to_convert_path)
    if not file_to_convert_path.name.endswith(tuple(RDF_FILE_ENDINGS.keys())):
        raise ValueError(
            "Files for conversion to Excel must end with one of the RDF file formats: '{}'".format(
                "', '".join(RDF_FILE_ENDINGS.keys())
            )
        )
    if profile not in profiles.PROFILES.keys():
        raise ValueError(
            "The profile chosen for conversion must be one of '{}'. 'vocpub' is default".format(
                "', '".join(profiles.PROFILES.keys())
            )
        )
    
    allow_warnings = True if error_level > 1 else False

    # validate the RDF file
    r = pyshacl.validate(
        str(file_to_convert_path), 
        shacl_graph=str(Path(__file__).parent / "validator.vocpub.ttl"), 
        allow_warnings=allow_warnings
    )
        
    logging_level = logging.INFO

    if message_level == 3:
        logging_level = logging.ERROR
    elif message_level == 2:
        logging_level = logging.WARNING

    if log_file:
        logging.basicConfig(level=logging_level, format="%(message)s", filename=log_file, force=True)
    else:
        logging.basicConfig(level=logging_level, format="%(message)s")

    info_list = []
    warning_list = []
    violation_list = []

    results_graph = r[1]
    from rdflib import Graph
    from rdflib.namespace import RDF, SH
    for report in results_graph.subjects(RDF.type, SH.ValidationReport):
        for result in results_graph.objects(report, SH.result):
            result_dict = {}
            for p, o in results_graph.predicate_objects(result):
                if p == SH.focusNode:
                    result_dict["focusNode"] = str(o)
                elif p == SH.resultMessage:
                    result_dict["resultMessage"] = str(o)
                elif p == SH.resultSeverity:
                    result_dict["resultSeverity"] = str(o)
                elif p == SH.sourceConstraintComponent:
                    result_dict["sourceConstraintComponent"] = str(o)
                elif p == SH.sourceShape:
                    result_dict["sourceShape"] = str(o)
                elif p == SH.value:
                    result_dict["value"] = str(o)
            result_message_formatted = log_msg(result_dict, log_file)
            result_message = log_msg(result_dict, "placeholder")
            if result_dict["resultSeverity"] == str(SH.Info):
                logging.info(result_message_formatted)
                info_list.append(result_message)
            elif result_dict["resultSeverity"] == str(SH.Warning):
                logging.warning(result_message_formatted)
                warning_list.append(result_message)
            elif result_dict["resultSeverity"] == str(SH.Violation):
                logging.error(result_message_formatted)
                violation_list.append(result_message)
    
    error_messages = []

    if error_level == 3:
        error_messages = violation_list
    elif error_level == 2:
        error_messages = warning_list + violation_list
    else: # error_level == 1
        error_messages = info_list + warning_list + violation_list
    
    if len(error_messages) > 0:
        raise ConversionError(
            f"The file you supplied is not valid according to the {profile} profile."
        )

    # the RDF is valid so extract data and create Excel
    from rdflib import Graph, Namespace, URIRef, Literal
    from rdflib.namespace import DCTERMS, PROV, RDF, RDFS, SKOS, OWL

    g = Graph().parse(
        str(file_to_convert_path), format=RDF_FILE_ENDINGS[file_to_convert_path.suffix]
    )

    wb = load_workbook(filename=(Path(__file__).parent / "blank.xlsx"))

    holder = {"hasTopConcept": [], "provenance": None}
    for s in g.subjects(RDF.type, SKOS.ConceptScheme):
        holder["uri"] = str(s)
        for p, o in g.predicate_objects(s):
            if p == SKOS.prefLabel:
                holder["title"] = o.toPython()
            elif p == SKOS.definition:
                holder["description"] = str(o)
            elif p == SKOS.note: ## AB
                holder["note"] = str(o)
                
            elif p == DCTERMS.created:
                holder["created"] = o.toPython()
            elif p == DCTERMS.modified:
                holder["modified"] = o.toPython()
            elif p == DCTERMS.creator:
                holder["creator"] = models.ORGANISATIONS_INVERSE[o]
            elif p == DCTERMS.publisher:
                holder["publisher"] = models.ORGANISATIONS_INVERSE[o]
            elif p == OWL.versionInfo:
                holder["versionInfo"] = str(o)
            elif p == DCTERMS.source:
                holder["provenance"] = str(o)
            elif p == DCTERMS.provenance:
                holder["provenance"] = str(o)
            elif p == PROV.wasDerivedFrom:
                holder["provenance"] = str(o)
            elif p == SKOS.hasTopConcept:
                holder["hasTopConcept"].append(str(o))

    # from models import ConceptScheme, Concept, Collection
    cs = models.ConceptScheme(
        uri=holder["uri"],
        title=holder["title"],
        description=holder["description"],
        created=holder["created"],
        modified=holder["modified"],
        creator=holder["creator"],
        publisher=holder["publisher"],
        version=holder["versionInfo"]
        if holder.get("versionInfo") is not None
        else None,
        provenance=holder["provenance"]
        if holder.get("provenance") is not None
        else None,
        custodian=None,
        pid=None,
    )
    cs.to_excel(wb)

    # infer inverses
    for s, o in g.subject_objects(SKOS.broader):
        g.add((o, SKOS.narrower, s))

    row_no = 16
    for s in g.subjects(RDF.type, SKOS.Concept):
        holder = {
            "uri": str(s),
            "children": [],
            "other_ids": [],
            "home_vocab_uri": None,
            "provenance": None,
        }
        for p, o in g.predicate_objects(s):
            if p == SKOS.prefLabel:
                holder["pref_label"] = o.toPython()
            elif p == SKOS.definition:
                holder["definition"] = str(o)
            elif p == SKOS.narrower:
                holder["children"].append(str(o))
            elif p == SKOS.notation:
                holder["other_ids"].append(str(o))
            elif p == RDFS.isDefinedBy:
                holder["home_vocab_uri"] = str(o)
            elif p == DCTERMS.source:
                holder["provenance"] = str(o)
            elif p == DCTERMS.provenance:
                holder["provenance"] = str(o)
            elif p == PROV.wasDerivedFrom:
                holder["provenance"] = str(o)

        models.Concept(
            uri=holder["uri"],
            pref_label=holder["pref_label"],
            definition=holder["definition"],
            children=holder["children"],
            other_ids=holder["other_ids"],
            home_vocab_uri=holder["home_vocab_uri"],
            provenance=holder["provenance"]
            if holder.get("provenance") is not None
            else None,
        ).to_excel(wb, row_no)
        row_no += 1

    row_no += 2

    for s in g.subjects(RDF.type, SKOS.Collection):
        holder = {
            "uri": str(s),
            "members": [],
        }
        for p, o in g.predicate_objects(s):
            if p == SKOS.prefLabel:
                holder["pref_label"] = o.toPython()
            elif p == SKOS.definition:
                holder["definition"] = str(o)
            elif p == SKOS.member:
                holder["members"].append(str(o))
            elif p == DCTERMS.source:
                holder["provenance"] = str(o)
            elif p == DCTERMS.provenance:
                holder["provenance"] = str(o)
            elif p == PROV.wasDerivedFrom:
                holder["provenance"] = str(o)

        models.Collection(
            uri=holder["uri"],
            pref_label=holder["pref_label"],
            definition=holder["definition"],
            members=holder["members"],
            provenance=holder["provenance"]
            if holder.get("provenance") is not None
            else None,
        ).to_excel(wb, row_no)
        row_no += 1

    if output_file_path is not None:
        dest = output_file_path
    else:
        dest = file_to_convert_path.with_suffix(".xlsx")
    wb.save(filename=dest)
    return dest

def log_msg(result: Dict, log_file: str) -> str:
    from rdflib.namespace import SH
    formatted_msg = ""
    message = f"""Validation Result in {result['sourceConstraintComponent'].split(str(SH))[1]} ({result['sourceConstraintComponent']}):
\tSeverity: sh:{result['resultSeverity'].split(str(SH))[1]}
\tSource Shape: <{result['sourceShape']}>
\tFocus Node: <{result['focusNode']}>
\tValue Node: <{result['value']}>
\tMessage: {result['resultMessage']}
"""
    if result["resultSeverity"] == str(SH.Info):
        formatted_msg = f"INFO: {message}" if log_file else Fore.BLUE + "INFO: " + Style.RESET_ALL + message
    elif result["resultSeverity"] == str(SH.Warning):
        formatted_msg = f"WARNING: {message}" if log_file else Fore.YELLOW + "WARNING: " + Style.RESET_ALL + message
    elif result["resultSeverity"] == str(SH.Violation):
        formatted_msg = f"VIOLATION: {message}" if log_file else Fore.RED + "VIOLATION: " + Style.RESET_ALL + message
    return formatted_msg


def main(args=None):
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )

    parser.add_argument(
        "-v",
        "--version",
        help="The version of this copy of VocExel. Must still set an file_to_convert value to call this (can be fake)",
        action="store_true",
    )

    parser.add_argument(
        "-lp",
        "--listprofiles",
        help="This flag, if set, must be the only flag supplied. It will cause the program to list all the vocabulary"
        " profiles that this converter, indicating both their URI and their short token for use with the"
        " -p (--profile) flag when converting Excel files",
        action="store_true",
    )

    parser.add_argument(
        "file_to_convert",
        type=Path,
        help="The Excel file to convert to a SKOS vocabulary in RDF or an RDF file to convert to an Excel file",
    )

    parser.add_argument(
        "-val", "--validate", help="Validate output file", action="store_true"
    )

    parser.add_argument(
        "-p",
        "--profile",
        help="A profile - a specified information model - for a vocabulary. This tool understands several profiles and"
        "you can choose which one you want to convert the Excel file according to. The list of profiles - URIs "
        "and their corresponding tokens - supported by VocExcel, can be found by running the program with the "
        "flag -lp or --listprofiles.",
        default="vocpub",
    )

    parser.add_argument(
        "-ot",
        "--outputtype",
        help="The format of the vocabulary output.",
        choices=["file", "string"],
        default="file",
    )

    parser.add_argument(
        "-o",
        "--outputfile",
        help="An optionally-provided output file path.",
        required=False,
    )

    parser.add_argument(
        "-of",
        "--outputformat",
        help="An optionally-provided output format for RDF files. Only relevant in Excel-to-RDf conversions.",
        required=False,
        choices=["turtle", "xml", "json-ld"],
        default="turtle",
    )

    parser.add_argument(
        "-s",
        "--sheet",
        help="The sheet within the target Excel Workbook to process",
        default="vocabulary",
    )

    # 1 - info, 2 - warning, 3 - violation
    # error severity level
    parser.add_argument(
        "-e",
        "--errorlevel",
        help="The minimum severity level which fails validation",
        default=1,
    )

    # print severity level
    parser.add_argument(
        "-m",
        "--messagelevel",
        help="The minimum severity level printed to console",
        default=1,
    )

    # log to file
    parser.add_argument(
        "-l",
        "--logfile",
        help="The file to write logging output to",
        required=False,
    )

    args = parser.parse_args()

    if args.listprofiles:
        s = "Profiles\nToken\tIRI\n-----\t-----\n"
        for k, v in profiles.PROFILES.items():
            s += f"{k}\t{v.uri}\n"

        print(s.rstrip())
        exit()
    elif args.version:
        print(__version__)
        exit()
    elif args.file_to_convert:
        if not args.file_to_convert.name.endswith(tuple(KNOWN_FILE_ENDINGS)):
            print(
                "Files for conversion must either end with .xlsx (Excel) or one of the known RDF file endings, '{}'"
                .format("', '".join(RDF_FILE_ENDINGS.keys()))
            )
            exit()

        print(f"Processing file {args.file_to_convert}")

        if args.file_to_convert.name.endswith(tuple(EXCEL_FILE_ENDINGS)):
            try:
                o = excel_to_rdf(
                    args.file_to_convert,
                    sheet_name=args.sheet,
                    output_type=args.outputtype,
                    output_file_path=args.outputfile,
                    output_format=args.outputformat,
                )
                if args.outputtype == "string":
                    print(o)
                else:
                    print(f"Output is file {o}")
            except Exception as e:
                print(e)
                exit()
        else:  # RDF file ending
            try:
                o = rdf_to_excel(
                    args.file_to_convert, 
                    profile=args.profile, 
                    output_file_path=args.outputfile, 
                    error_level=int(args.errorlevel), 
                    message_level=int(args.messagelevel), 
                    log_file=args.logfile
                )
                if args.outputtype == "string":
                    print(o)
                else:
                    print(f"Output is file {o}")
            except Exception as e:
                print(e)
                exit()


if __name__ == "__main__":
    import sys

    main(sys.argv[1:])
