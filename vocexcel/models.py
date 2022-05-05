import datetime
from typing import List

from openpyxl import Workbook
from pydantic import AnyHttpUrl
from pydantic import BaseModel, validator
from rdflib import Graph, URIRef, Literal
from rdflib.namespace import DCAT, DCTERMS, OWL, SKOS, RDF, RDFS, XSD

ORGANISATIONS = {
    "LIKAT": URIRef("https://www.catalysis.de/home"),
    "TUDO": URIRef("https://www.tu-dortmund.de/"),
    "UHGW": URIRef("https://www.uni-greifswald.de/"),
    "NFDI4Cat": URIRef("https://nfdi4cat.org/"),
    "FAU": URIRef("https://www.fau.de/"),
    "FOKUS": URIRef("https://www.fokus.fraunhofer.de/"),
    "HLRS": URIRef("https://www.hlrs.de/"),
    "KIT": URIRef("https://www.kit.edu/"),
    "MPI CEC": URIRef("https://www.cec.mpg.de/"),
    "MPI DCTS": URIRef("https://www.mpi-magdeburg.mpg.de/"),
    "RWTH": URIRef("https://www.rwth-aachen.de/"),
    "TUB": URIRef("https://www.tu.berlin/"),
    "TUBS": URIRef("https://www.tu-braunschweig.de/"),
    "TUM": URIRef("https://www.tum.de/"),
    "UHRO": URIRef("https://www.uni-rostock.de/"),
    "UL": URIRef("https://www.uni-leipzig.de/"),

}

ORGANISATIONS_INVERSE = {
    URIRef("https://www.catalysis.de/home"): "LIKAT",
    URIRef("https://www.tu-dortmund.de/"): "TUDO",
    URIRef("https://www.uni-greifswald.de/"): "UHGW",
    URIRef("https://nfdi4cat.org/"): "NFDI4Cat",
    URIRef("https://www.fau.de/"): "FAU",
    URIRef("https://www.fokus.fraunhofer.de/"): "FOKUS",
    URIRef("https://www.hlrs.de/"): "HLRS",
    URIRef("https://www.kit.edu/"): "KIT",
    URIRef("https://www.cec.mpg.de/"): "MPI CEC",
    URIRef("https://www.mpi-magdeburg.mpg.de/"): "MPI DCTS",
    URIRef("https://www.rwth-aachen.de/"): "RWTH",
    URIRef("https://www.tu.berlin/"): "TUB",
    URIRef("https://www.tu-braunschweig.de/"): "TUBS",
    URIRef("https://www.tum.de/"): "TUM",
    URIRef("https://www.uni-rostock.de/"): "UHRO",
    URIRef("https://www.uni-leipzig.de/"): "UL",
}


class ConceptScheme(BaseModel):
    uri: AnyHttpUrl
    title: str
    description: str
    created: datetime.date
    modified: datetime.date = None
    creator: str
    publisher: str
    provenance: str
    version: str = None
    custodian: str = None
    pid: AnyHttpUrl = None

    @validator("creator")
    def creator_must_be_from_list(cls, v):
        if v not in ORGANISATIONS.keys():
            raise ValueError(
                f"Organisations must selected from the Organisations list: {', '.join(ORGANISATIONS)}"
            )
        return v

    @validator("publisher")
    def publisher_must_be_from_list(cls, v):
        if v not in ORGANISATIONS.keys():
            raise ValueError(
                f"Organisations must selected from the Organisations list: {', '.join(ORGANISATIONS)}"
            )
        return v

    def to_graph(self):
        g = Graph()
        v = URIRef(self.uri)
        g.add((v, RDF.type, SKOS.ConceptScheme))
        g.add((v, SKOS.prefLabel, Literal(self.title, lang="en")))
        g.add((v, SKOS.definition, Literal(self.description, lang="en")))
        g.add((v, DCTERMS.created, Literal(self.created, datatype=XSD.date)))
        if self.modified is not None:
            g.add((v, DCTERMS.modified, Literal(self.created, datatype=XSD.date)))
        else:
            g.add(
                (
                    v,
                    DCTERMS.modified,
                    Literal(
                        datetime.datetime.now().strftime("%Y-%m-%d"), datatype=XSD.date
                    ),
                )
            )
        g.add((v, DCTERMS.creator, ORGANISATIONS[self.creator]))
        g.add((v, DCTERMS.publisher, ORGANISATIONS[self.publisher]))
        if self.version is not None:
            g.add((v, OWL.versionInfo, Literal(self.version)))
        g.add((v, DCTERMS.provenance, Literal(self.provenance, lang="en")))
        if self.custodian is not None:
            g.add((v, DCAT.contactPoint, Literal(self.custodian)))
        if self.pid is not None:
            g.add((v, RDFS.seeAlso, URIRef(self.pid)))

        # bind non-core prefixes
        g.bind("cs", v)
        g.bind(
            "",
            str(v).split("#")[0] if "#" in str(v) else "/".join(str(v).split("/")[:-1]),
        )
        g.bind("dcat", DCAT)
        g.bind("dcterms", DCTERMS)
        g.bind("skos", SKOS)
        g.bind("owl", OWL)

        return g

    def to_excel(self, wb: Workbook):
        ws = wb.active
        ws["B1"] = self.uri
        ws["B2"] = self.title
        ws["B3"] = self.description
        ws["B4"] = self.created.isoformat()
        ws["B5"] = self.modified.isoformat()
        ws["B6"] = self.creator
        ws["B7"] = self.publisher
        ws["B8"] = self.version
        ws["B9"] = self.provenance
        # ws["B10"] = ""
        # ws["B11"] = ""


class Concept(BaseModel):
    uri: str
    pref_label: str
    alt_labels: List[str] = None
    pl_language_code: List[str] = None
    definition: str
    def_language_code: List[str] = None
    children: List[str] = None
    related: List[str] = None
    other_ids: List[str] = None
    home_vocab_uri: str = None
    provenance: str = None

    def to_graph(self):
        g = Graph()
        c = URIRef(self.uri)
        g.add((c, RDF.type, SKOS.Concept))
        if self.pl_language_code is None:
            self.pl_language_code = ["en"]
        for lang_code in self.pl_language_code:
            g.add((c, SKOS.prefLabel, Literal(self.pref_label, lang=lang_code)))
        if self.alt_labels is not None:
            for alt_label in self.alt_labels:
                g.add((c, SKOS.altLabel, Literal(alt_label, lang="en")))
        if self.def_language_code is None:
            self.def_language_code = ["en"]
        for lang_code in self.def_language_code:
            g.add((c, SKOS.definition, Literal(self.definition, lang=lang_code)))
        if self.children is not None:
            for child in self.children:
                g.add((c, SKOS.narrower, URIRef(child)))
                g.add((URIRef(child), SKOS.broader, c))
        if self.related is not None:
            for relation in self.related:
                g.add((c, SKOS.related, URIRef(relation)))
                g.add((URIRef(relation), SKOS.related, c))
        if self.other_ids is not None:
            for other_id in self.other_ids:
                g.add((c, SKOS.notation, Literal(other_id)))
        if self.home_vocab_uri is not None:
            g.add((c, RDFS.isDefinedBy, URIRef(self.home_vocab_uri)))
        if self.provenance is not None:
            g.add((c, DCTERMS.provenance, Literal(self.provenance, lang="en")))

        return g

    def to_excel(self, wb: Workbook, row_no: int):
        ws = wb.active
        ws[f"A{row_no}"] = self.uri
        ws[f"B{row_no}"] = self.pref_label
        ws[f"C{row_no}"] = self.alt_labels
        ws[f"D{row_no}"] = self.definition
        ws[f"E{row_no}"] = ",\\".join(self.children)
        ws[f"F{row_no}"] = ",\\".join(self.related)
        ws[f"G{row_no}"] = ",\\".join(self.other_ids)
        ws[f"H{row_no}"] = self.home_vocab_uri
        ws[f"I{row_no}"] = self.provenance


class Collection(BaseModel):
    uri: str
    pref_label: str
    definition: str
    members: List[str]
    provenance: str = None

    @validator("members")
    def members_must_by_iris(cls, v):
        if not v[0].startswith("http"):
            raise ValueError("The members of a Collection must be a list of IRIs")

    def to_graph(self):
        g = Graph()
        c = URIRef(self.uri)
        g.add((c, RDF.type, SKOS.Collection))
        g.add((c, SKOS.prefLabel, Literal(self.pref_label, lang="en")))
        g.add((c, SKOS.definition, Literal(self.definition, lang="en")))
        for member in self.members:
            g.add((c, SKOS.member, URIRef(member)))
        if self.provenance is not None:
            g.add((c, DCTERMS.provenance, Literal(self.provenance, lang="en")))

        return g

    def to_excel(self, wb: Workbook, row_no: int):
        ws = wb.active
        ws[f"A{row_no}"] = self.uri
        ws[f"B{row_no}"] = self.pref_label
        ws[f"C{row_no}"] = self.definition
        ws[f"D{row_no}"] = ",\n".join(self.members)
        ws[f"E{row_no}"] = self.provenance


class Vocabulary(BaseModel):
    concept_scheme: ConceptScheme
    concepts: List[Concept]
    collections: List[Collection]

    def to_graph(self):
        g = self.concept_scheme.to_graph()
        cs = URIRef(self.concept_scheme.uri)
        for concept in self.concepts:
            g += concept.to_graph()
            g.add((URIRef(concept.uri), SKOS.inScheme, cs))
        for collection in self.collections:
            g += collection.to_graph()
            g.add((URIRef(collection.uri), DCTERMS.isPartOf, cs))
            g.add((cs, DCTERMS.hasPart, URIRef(collection.uri)))

        # create as Top Concepts those Concepts that have no skos:narrower properties with them as objects
        for s in g.subjects(SKOS.inScheme, cs):
            is_tc = True
            for _ in g.objects(s, SKOS.broader):
                is_tc = False
            if is_tc:
                g.add((cs, SKOS.hasTopConcept, s))
                g.add((s, SKOS.topConceptOf, cs))

        return g
